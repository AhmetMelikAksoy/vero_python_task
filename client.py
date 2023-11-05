import argparse
import requests
import pandas as pd
import datetime
import json
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

SERVER_URL = "http://0.0.0.0:8000"

# logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logging.basicConfig(filename='logs/client.log', level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

def main():
    parser = argparse.ArgumentParser(description="Client for interacting with the REST API server.")
    parser.add_argument("-k", "--keys", nargs="+", help="Additional keys to consider as columns")
    parser.add_argument("-c", "--colored", action="store_true", help="Enable row coloring")
    args = parser.parse_args()

    file_path = "vehicles.csv"

    df = pd.read_csv(file_path, sep=";")
    json_data = df.to_json(orient='split')
    payload = {'data': json_data}

    response = requests.post("http://localhost:8000/process-vehicles/", json=payload)

    if response.status_code == 200:
        logger.info("Successful request, new data recieved.")
        result = response.json()
        data = json.loads(result)
        df_server = pd.DataFrame(data)

        possible_columns = set(df_server.columns)
        columns_to_include = ['rnr', 'gruppe', 'hu']

        if args.keys is not None and 'labelIds' in args.keys:
            for key in args.keys:
                if key in possible_columns and key not in columns_to_include:
                    columns_to_include.append(key)

        df_server = df_server.sort_values(by='gruppe')
        current_date_iso_formatted = datetime.datetime.now().strftime("%Y-%m-%d")
        excel_file_name = f'vehicles_{current_date_iso_formatted}.xlsx'
        writer = pd.ExcelWriter(excel_file_name, engine='openpyxl')
        df_server_keys = df_server[columns_to_include]
        df_server_keys.to_excel(writer, sheet_name='Vehicles', index=False)

        
        if args.colored:
            hu_colors = []
            for _, row in df_server_keys.iterrows():
                hu = row['hu']
                color = ""
                if pd.notna(hu):
                    hu_date = pd.to_datetime(hu)
                    today = pd.to_datetime(datetime.date.today())
                    delta = today - hu_date
                    if delta.days <= 90:
                        color = "007500"  # green
                    elif delta.days <= 365:
                        color = "FFA500"  # orange
                    else:
                        color = "b30000"  # red
                hu_colors.append(color)

            current_date_iso_formatted = datetime.datetime.now().strftime("%Y-%m-%d")
            excel_file_name = f'vehicles_{current_date_iso_formatted}.xlsx'

            wb = openpyxl.Workbook()
            ws = wb.active

            for r_idx, row in enumerate(dataframe_to_rows(df_server_keys, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

            for row_num, color in enumerate(hu_colors, start=2):
                if color:
                    for col_num in range(1, len(df_server_keys.columns) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.fill = fill

            if args.keys is not None and 'labelIds' in args.keys:
                for _, row in df_server.iterrows():
                    label_id = row.get('labelIds')
                    if not pd.isna(label_id):
                        color_code = row["resolved_colorCode"]
                        labelIds_col_num = df_server_keys.columns.get_loc("labelIds") + 1 #excel sheet column number of labelIds
                        for row_num in range(2,df_server_keys.shape[0]):
                            cell = ws.cell(row=row_num, column=labelIds_col_num)
                            if color_code:
                                aRGB_color_code = "FF" + color_code.lstrip('#') #aRGB color code needed
                                cell.font = openpyxl.styles.Font(color=aRGB_color_code)

            wb.save(excel_file_name)
            logger.info(f"Excel file '{excel_file_name}' with color-coding created.")
        else:
            df_server_keys.to_excel(excel_file_name, index=False, engine='openpyxl')
            logger.info(f"Excel file '{excel_file_name}' created.")

    else:
        logger.error(f"Error: {response.status_code} - {response.text}")


if __name__ == "__main__":
    main()